from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
import subprocess
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
WULIU_DIR = PROJECT_ROOT / "wuliu"
DEFAULT_TESSDATA_PREFIX = PROJECT_ROOT.parent / "tessdata"
STRUCTURED_CONTRACT_ID = "kjds-ru002-logistics-observation-v1"

COST_LEG_KEYWORDS = {
    "domestic_logistics": ("内地快递", "国内快递", "退回内地"),
    "international_logistics": (
        "运费",
        "OZON",
        "Yandex",
        "WB",
        "PUDO",
        "Courier",
        "FBP",
        "rFBS",
    ),
    "packaging": ("包装", "包材", "气泡", "快递袋", "珍珠棉", "缠绕膜"),
    "warehousing": ("仓储", "仓库", "入库", "出库"),
    "customs": ("报关", "清关", "海关"),
    "last_mile": ("到门", "派送", "取货点", "末端"),
    "return": ("退货", "退回", "退仓", "退件", "销毁", "拦截"),
    "customer_compensation": ("理赔", "赔付", "赔偿"),
    "damage": ("丢件", "破损", "销毁", "损失"),
}
COST_LEG_ORDER = tuple(COST_LEG_KEYWORDS)

KEYWORDS = (
    "运费",
    "价格",
    "费率",
    "计费",
    "时效",
    "仓库",
    "仓配",
    "PUDO",
    "Courier",
    "FBS",
    "FBP",
    "rFBS",
    "OZON",
    "Yandex",
    "WB",
    "退货",
    "销毁",
    "贴标",
    "标签",
    "保险",
    "报关",
    "仓储",
    "拍照",
    "称重",
    "拆单",
    "合包",
    "代贴",
    "清关",
)

FEE_PATTERN = re.compile(
    r"(?:\d+(?:[.,]\d+)?(?:\s*元|\s*RMB|\s*rub|\s*卢布|\s*РУБ))|(?:\d+(?:[.,]\d+)?\s*/\s*(?:kg|KG|千克|票|单|件|100g|100克|克))",
    re.IGNORECASE,
)
PHONE_PATTERN = re.compile(r"(?<!\d)(?:1\d{10}|\+?\d{2,4}[-\s]?\d{6,13})(?!\d)")
EMAIL_PATTERN = re.compile(r"(?i)(?<![\w.+-])[\w.+-]+@[\w.-]+\.[a-z]{2,}(?![\w.-])")


@dataclass(frozen=True, slots=True)
class EvidenceHit:
    source_relpath: str
    sha256: str
    kind: str
    location: str
    excerpt: str
    currency: str = "UNKNOWN"
    tax_treatment: str = "UNKNOWN"
    validity: str = "UNKNOWN"
    mapped_cost_leg: str = "UNKNOWN"
    mapped_cost_legs: tuple[str, ...] = ()
    status: str = "observed"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sanitize_text(text: str) -> str:
    text = PHONE_PATTERN.sub("[REDACTED_PHONE]", text)
    text = EMAIL_PATTERN.sub("[REDACTED_EMAIL]", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def source_relpath(path: Path) -> str:
    try:
        return path.relative_to(PROJECT_ROOT).as_posix()
    except ValueError:
        return f"external/{path.name}"


def _infer_currency(text: str) -> str:
    if "卢布" in text or "RUB" in text.upper() or "руб" in text.lower():
        return "RUB"
    if "元" in text or "RMB" in text.upper():
        return "CNY"
    return "UNKNOWN"


def _matches_interest(text: str) -> bool:
    if any(keyword in text for keyword in KEYWORDS):
        return True
    return bool(FEE_PATTERN.search(text))


def infer_cost_legs(text: str) -> tuple[str, ...]:
    normalized = text.upper()
    return tuple(
        cost_leg
        for cost_leg in COST_LEG_ORDER
        if any(keyword.upper() in normalized for keyword in COST_LEG_KEYWORDS[cost_leg])
    )


def structured_records(hits: Iterable[EvidenceHit]) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    for hit in hits:
        sanitized_source_relpath = sanitize_text(hit.source_relpath)
        sanitized_source_location = sanitize_text(hit.location)
        sanitized_excerpt = sanitize_text(hit.excerpt)[:1600]
        mapped_cost_legs = hit.mapped_cost_legs or infer_cost_legs(
            sanitized_excerpt
        )
        identity = {
            "contract_id": STRUCTURED_CONTRACT_ID,
            "source_relpath": sanitized_source_relpath,
            "source_sha256": hit.sha256,
            "source_location": sanitized_source_location,
            "excerpt": sanitized_excerpt,
        }
        identity_sha256 = hashlib.sha256(
            json.dumps(
                identity,
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            ).encode()
        ).hexdigest()
        record: dict[str, object] = {
            **identity,
            "observation_id": f"ru002_{identity_sha256[:24]}",
            "source_kind": hit.kind,
            "source_excerpt_sanitized": True,
            "currency": hit.currency,
            "tax_treatment": hit.tax_treatment,
            "validity": hit.validity,
            "mapped_cost_legs": list(mapped_cost_legs),
            "evidence_level": "observed",
            "sku_binding": None,
            "variant_binding": None,
            "quantity_binding": None,
            "shipment_profile_binding": None,
            "effective_period": None,
            "decision_eligible": False,
            "actual_cost_created": False,
            "external_write_allowed": False,
            "status": hit.status,
        }
        record["observation_sha256"] = hashlib.sha256(
            json.dumps(
                record,
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            ).encode()
        ).hexdigest()
        records.append(record)
    return records


def extract_xlsx_hits(path: Path, *, row_limit: int = 30, col_limit: int = 24) -> list[EvidenceHit]:
    sha256 = sha256_file(path)
    workbook = load_workbook(path, data_only=False)
    hits: list[EvidenceHit] = []

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=min(row_limit, worksheet.max_row),
            min_col=1,
            max_col=min(col_limit, worksheet.max_column),
        ):
            row_values: list[str] = []
            row_contains_interest = False
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                text = sanitize_text(str(value))
                row_values.append(f"{cell.coordinate}={text}")
                row_contains_interest = row_contains_interest or _matches_interest(text)
            if row_values and row_contains_interest:
                joined = " | ".join(row_values)
                hits.append(
                    EvidenceHit(
                        source_relpath=source_relpath(path),
                        sha256=sha256,
                        kind="xlsx",
                        location=f"{sheet_name}!row:{row[0].row}",
                        excerpt=joined,
                        currency=_infer_currency(joined),
                    )
                )
    return hits


def _pypdf_reader():
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception:
        return None
    return PdfReader


def extract_pdf_hits(path: Path, *, page_limit: int = 8) -> list[EvidenceHit]:
    sha256 = sha256_file(path)
    PdfReader = _pypdf_reader()
    if PdfReader is None:
        return [
            EvidenceHit(
                source_relpath=source_relpath(path),
                sha256=sha256,
                kind="pdf",
                location="pdf",
                excerpt="pypdf unavailable; page text not extracted in base environment",
                status="unparsed_optional_dependency_missing",
            )
        ]

    reader = PdfReader(str(path))
    hits: list[EvidenceHit] = []
    for index, page in enumerate(reader.pages[:page_limit], start=1):
        text = sanitize_text(page.extract_text() or "")
        if not text:
            continue
        if _matches_interest(text):
            hits.append(
                EvidenceHit(
                    source_relpath=source_relpath(path),
                    sha256=sha256,
                    kind="pdf",
                    location=f"page:{index}",
                    excerpt=text[:1600],
                    currency=_infer_currency(text),
                )
            )
    if not hits:
        hits.append(
            EvidenceHit(
                source_relpath=source_relpath(path),
                sha256=sha256,
                kind="pdf",
                location="pdf",
                excerpt="pdf parsed but no fee-bearing text matched the interest filter",
                status="no_fee_hits",
            )
        )
    return hits


def _tesseract_exe() -> str:
    return shutil.which("tesseract") or r"C:\Program Files\Tesseract-OCR\tesseract.exe"


def extract_image_hits(path: Path, *, max_chars: int = 4000) -> list[EvidenceHit]:
    sha256 = sha256_file(path)
    exe = _tesseract_exe()
    if not Path(exe).exists():
        return [
            EvidenceHit(
                source_relpath=source_relpath(path),
                sha256=sha256,
                kind="image",
                location="image",
                excerpt="tesseract executable not found",
                status="unparsed_tesseract_missing",
            )
        ]

    env = os.environ.copy()
    env.setdefault("TESSDATA_PREFIX", str(DEFAULT_TESSDATA_PREFIX))
    completed = subprocess.run(
        [exe, str(path), "stdout", "-l", "chi_sim+eng"],
        capture_output=True,
        text=True,
        check=False,
        env=env,
    )
    ocr_text = sanitize_text((completed.stdout or completed.stderr or "")[:max_chars])
    if not ocr_text:
        return [
            EvidenceHit(
                source_relpath=source_relpath(path),
                sha256=sha256,
                kind="image",
                location="image",
                excerpt="OCR returned no text",
                status="no_ocr_text",
            )
        ]

    hits: list[EvidenceHit] = []
    for line_number, line in enumerate(ocr_text.splitlines(), start=1):
        if _matches_interest(line):
            hits.append(
                EvidenceHit(
                    source_relpath=source_relpath(path),
                    sha256=sha256,
                    kind="image",
                    location=f"ocr_line:{line_number}",
                    excerpt=line.strip(),
                    currency=_infer_currency(line),
                )
            )
    if not hits:
        hits.append(
            EvidenceHit(
                source_relpath=source_relpath(path),
                sha256=sha256,
                kind="image",
                location="image",
                excerpt=ocr_text[:1600],
                status="ocr_text_no_fee_hits",
            )
        )
    return hits


def extract_legacy_doc(path: Path) -> list[EvidenceHit]:
    sha256 = sha256_file(path)
    return [
        EvidenceHit(
            source_relpath=source_relpath(path),
            sha256=sha256,
            kind="doc",
            location="document",
            excerpt="legacy .doc file retained as source-of-truth evidence; structure not parsed in base environment",
            status="unsupported_legacy_doc",
        )
    ]


def scan_source(path: Path) -> list[EvidenceHit]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return extract_xlsx_hits(path)
    if suffix == ".pdf":
        return extract_pdf_hits(path)
    if suffix in {".jpg", ".jpeg", ".png"}:
        return extract_image_hits(path)
    if suffix == ".doc":
        return extract_legacy_doc(path)
    return [
        EvidenceHit(
            source_relpath=source_relpath(path),
            sha256=sha256_file(path),
            kind=suffix.lstrip(".") or "unknown",
            location="file",
            excerpt="unsupported source type",
            status="unsupported_source_type",
        )
    ]


def scan_wuliu_directory(root: Path = WULIU_DIR) -> list[EvidenceHit]:
    hits: list[EvidenceHit] = []
    for path in sorted(root.iterdir(), key=lambda item: item.name):
        if not path.is_file():
            continue
        if path.suffix.lower() not in {".xlsx", ".xlsm", ".pdf", ".jpg", ".jpeg", ".png", ".doc"}:
            continue
        hits.extend(scan_source(path))
    return hits


def render_markdown(hits: Iterable[EvidenceHit]) -> str:
    grouped: dict[str, list[EvidenceHit]] = {}
    for hit in hits:
        grouped.setdefault(hit.source_relpath, []).append(hit)

    lines: list[str] = [
        "# RU-002 Logistics Evidence Scan",
        "",
        "This report is generated from read-only inspection of `wuliu/` sources.",
        "Sensitive contact strings are redacted in the parser before the output is rendered.",
        "",
    ]
    for source, source_hits in grouped.items():
        first = source_hits[0]
        lines.append(f"## `{source}`")
        lines.append("")
        lines.append(f"- SHA-256: `{first.sha256}`")
        for hit in source_hits:
            lines.append(f"- `{hit.location}` [{hit.kind}] {hit.excerpt}")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def main() -> int:
    parser = argparse.ArgumentParser(description="Scan RU-002 logistics evidence sources read-only.")
    parser.add_argument("--json", action="store_true", help="Emit JSON instead of markdown.")
    parser.add_argument("--source", type=Path, default=WULIU_DIR, help="Directory to scan.")
    parser.add_argument("--output", type=Path, help="Write the rendered result to this path.")
    args = parser.parse_args()

    hits = scan_wuliu_directory(args.source)
    rendered = (
        json.dumps(structured_records(hits), ensure_ascii=False, indent=2)
        if args.json
        else render_markdown(hits)
    )
    try:
        if args.output:
            args.output.parent.mkdir(parents=True, exist_ok=True)
            args.output.write_text(rendered.rstrip() + "\n", encoding="utf-8")
        else:
            print(rendered)
    except BrokenPipeError:
        return 0
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

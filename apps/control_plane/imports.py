from __future__ import annotations

import csv
import hashlib
import io
from collections.abc import Iterable
from dataclasses import asdict, dataclass
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import (
    JSON,
    CheckConstraint,
    DateTime,
    ForeignKey,
    Index,
    Integer,
    String,
    UniqueConstraint,
    select,
    text,
)
from sqlalchemy.orm import Mapped, Session, mapped_column

from .domain import new_id
from .evidence import EvidenceRecordRow
from .ozon_contracts import CONTRACTS, OzonRecordType, detect_record_type, normalize_record
from .sql_repository import Base

MAX_IMPORT_BYTES = 20 * 1024 * 1024
MAX_IMPORT_ROWS = 50_000

FIELD_ALIASES = {
    "store_ref": {"store_ref", "store", "shop_ref", "店铺"},
    "external_id": {
        "external_id",
        "operation_id",
        "document_id",
        "order_id",
        "order_number",
        "номер документа",
        "номер заказа",
        "номер операции",
        "номер отправления",
        "номер возврата",
        "id начисления",
    },
    "order_external_id": {
        "order_external_id",
        "parent_order_id",
        "original_order_id",
        "номер исходного заказа",
        "номер заказа возврата",
    },
    "sku": {"sku", "offer_id", "артикул", "артикул продавца"},
    "quantity": {"quantity", "qty", "количество"},
    "currency": {"currency", "валюта"},
    "gross_revenue": {"gross_revenue", "price", "revenue", "сумма", "цена", "итого"},
    "status": {"status", "статус"},
    "effective_at": {
        "effective_at",
        "created_at",
        "order_date",
        "date",
        "дата",
        "дата заказа",
        "дата операции",
        "дата начисления",
        "дата выплаты",
    },
    "fee_type": {"fee_type", "service", "commission_type", "услуга", "тип начисления"},
    "accrual_group": {"accrual_group", "группа услуг"},
    "accrual_type": {"accrual_type", "тип начисления"},
    "amount": {
        "amount",
        "fee_amount",
        "settlement_amount",
        "refund_amount",
        "итого",
        "итого, руб",
        "сумма выплаты",
        "сумма операции",
        "сумма итого, руб.",
    },
    "return_reason": {"return_reason", "reason", "причина возврата"},
    "warehouse_ref": {
        "warehouse_ref",
        "warehouse_id",
        "warehouse",
        "склад",
        "идентификатор склада",
    },
    "cluster_ref": {
        "cluster_ref",
        "cluster",
        "region_cluster",
        "кластер",
    },
    "fulfillment_mode": {
        "fulfillment_mode",
        "delivery_scheme",
        "scheme",
        "схема работы",
        "схема доставки",
    },
    "available_quantity": {
        "available_quantity",
        "available_stock",
        "present",
        "available",
        "доступно",
        "доступный остаток",
    },
    "reserved_quantity": {
        "reserved_quantity",
        "reserved_stock",
        "reserved",
        "зарезервировано",
        "резерв",
    },
    "in_transit_quantity": {
        "in_transit_quantity",
        "in_transit",
        "inbound",
        "в пути",
    },
    "damaged_quantity": {
        "damaged_quantity",
        "damaged",
        "брак",
        "повреждено",
    },
    "quarantine_quantity": {
        "quarantine_quantity",
        "quarantine",
        "карантин",
    },
}

DERIVED_RUB_FROM_HEADER = "__derived_rub_from_amount_header__"


class ImportJobRow(Base):
    __tablename__ = "import_jobs"
    __table_args__ = (
        CheckConstraint(
            "("
            "tenant_ref IS NULL AND entity_ref IS NULL "
            "AND store_ref IS NULL "
            "AND scope_grant_authority_sha256 IS NULL "
            "AND source_evidence_sha256 IS NULL "
            "AND scope_as_of IS NULL"
            ") OR ("
            "tenant_ref IS NOT NULL AND length(tenant_ref) > 0 "
            "AND entity_ref IS NOT NULL AND length(entity_ref) > 0 "
            "AND store_ref IS NOT NULL AND length(store_ref) > 0 "
            "AND scope_grant_authority_sha256 IS NOT NULL "
            "AND length(scope_grant_authority_sha256) = 64 "
            "AND source_evidence_sha256 IS NOT NULL "
            "AND length(source_evidence_sha256) = 64 "
            "AND scope_as_of IS NOT NULL "
            "AND evidence_id IS NOT NULL"
            ")",
            name="ck_import_job_scope_complete",
        ),
        Index(
            "uq_import_job_legacy_sha256",
            "sha256",
            unique=True,
            sqlite_where=text("tenant_ref IS NULL"),
            postgresql_where=text("tenant_ref IS NULL"),
        ),
        Index(
            "uq_import_job_scoped_sha256",
            "tenant_ref",
            "entity_ref",
            "store_ref",
            "sha256",
            unique=True,
            sqlite_where=text("tenant_ref IS NOT NULL"),
            postgresql_where=text("tenant_ref IS NOT NULL"),
        ),
        Index(
            "ix_import_job_scope_created",
            "tenant_ref",
            "entity_ref",
            "store_ref",
            "created_at",
        ),
    )
    id: Mapped[str] = mapped_column(String, primary_key=True)
    source: Mapped[str] = mapped_column(String)
    record_type: Mapped[str] = mapped_column(String)
    filename: Mapped[str] = mapped_column(String)
    sha256: Mapped[str] = mapped_column(String(64))
    status: Mapped[str] = mapped_column(String)
    row_count: Mapped[int] = mapped_column(Integer)
    accepted_count: Mapped[int] = mapped_column(Integer)
    rejected_count: Mapped[int] = mapped_column(Integer)
    mapping_json: Mapped[dict[str, str]] = mapped_column(JSON)
    errors_json: Mapped[list[dict[str, Any]]] = mapped_column(JSON)
    evidence_id: Mapped[str | None] = mapped_column(ForeignKey(EvidenceRecordRow.id), nullable=True)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True))
    tenant_ref: Mapped[str | None] = mapped_column(String(160), nullable=True)
    entity_ref: Mapped[str | None] = mapped_column(String(160), nullable=True)
    store_ref: Mapped[str | None] = mapped_column(String(160), nullable=True)
    scope_grant_authority_sha256: Mapped[str | None] = mapped_column(
        String(64),
        nullable=True,
    )
    source_evidence_sha256: Mapped[str | None] = mapped_column(
        String(64),
        nullable=True,
    )
    scope_as_of: Mapped[datetime | None] = mapped_column(
        DateTime(timezone=True),
        nullable=True,
    )


class ImportDataRow(Base):
    __tablename__ = "import_rows"
    __table_args__ = (UniqueConstraint("import_id", "row_number"),)
    id: Mapped[str] = mapped_column(String, primary_key=True)
    import_id: Mapped[str] = mapped_column(ForeignKey("import_jobs.id"))
    row_number: Mapped[int] = mapped_column(Integer)
    record_type: Mapped[str] = mapped_column(String)
    external_id: Mapped[str | None] = mapped_column(String, nullable=True)
    payload_json: Mapped[dict[str, Any]] = mapped_column(JSON)
    normalized_json: Mapped[dict[str, Any]] = mapped_column(JSON)
    errors_json: Mapped[list[str]] = mapped_column(JSON)


@dataclass(frozen=True, slots=True)
class ImportResult:
    id: str
    source: str
    record_type: str
    filename: str
    sha256: str
    status: str
    row_count: int
    accepted_count: int
    rejected_count: int
    mapping: dict[str, str]
    errors: list[dict[str, Any]]
    evidence_id: str | None
    created_at: str
    duplicate: bool = False
    scope: dict[str, Any] | None = None
    formal_fact_promotion_allowed: bool = False
    external_write_allowed: bool = False

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True, slots=True)
class ImportPreview:
    filename: str
    sha256: str
    record_type: str
    row_count: int
    mapping: dict[str, str]
    missing_columns: list[str]
    ready: bool


def _clean_header(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _detect_mapping(headers: list[str]) -> dict[str, str]:
    normalized_headers = {_clean_header(header): header for header in headers}
    mapping: dict[str, str] = {}
    for canonical, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if alias in normalized_headers:
                mapping[canonical] = normalized_headers[alias]
                break
    if "сумма итого, руб." in normalized_headers and "currency" not in mapping:
        mapping["currency"] = DERIVED_RUB_FROM_HEADER
    return mapping


def _normalize(
    raw: dict[str, Any],
    mapping: dict[str, str],
    record_type: OzonRecordType,
    *,
    source_row_id: str | None = None,
) -> tuple[dict[str, Any], list[str]]:
    result: dict[str, Any] = {}
    for canonical, header in mapping.items():
        if header == DERIVED_RUB_FROM_HEADER:
            result[canonical] = "RUB"
            continue
        value = raw.get(header)
        if canonical == "effective_at" and isinstance(value, datetime) and value.tzinfo is None:
            value = value.replace(tzinfo=UTC)
        result[canonical] = str("" if value is None else value).strip()
    if record_type is OzonRecordType.ACCRUAL and not result.get("external_id") and source_row_id:
        result["external_id"] = source_row_id
    return normalize_record(record_type, result)


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, Decimal)):
        return str(value)
    return str(value)


def _csv_rows(content: bytes) -> Iterable[dict[str, Any]]:
    decoded: str | None = None
    for encoding in ("utf-8-sig", "cp1251"):
        try:
            decoded = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise ValueError("CSV encoding must be UTF-8 or Windows-1251")
    sample = decoded[:4096]
    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    yield from csv.DictReader(io.StringIO(decoded), dialect=dialect)


def _xlsx_rows(content: bytes) -> Iterable[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    values = sheet.iter_rows(values_only=True)
    header_row = next(values, ())
    nonempty_header_cells = [value for value in header_row if str(value or "").strip()]
    if len(nonempty_header_cells) == 1 and str(nonempty_header_cells[0]).strip().lower().startswith("период:"):
        header_row = next(values, ())
    headers = [str(value or "").strip() for value in header_row]
    for row in values:
        yield dict(zip(headers, row, strict=False))


class OzonImportService:
    def __init__(self, engine) -> None:
        self.engine = engine

    def find_by_content(
        self,
        content: bytes,
        *,
        scope_authority: dict[str, Any] | None = None,
    ) -> ImportResult | None:
        digest = hashlib.sha256(content).hexdigest()
        scope = self._scope_authority(scope_authority)
        query = select(ImportJobRow).where(ImportJobRow.sha256 == digest)
        if scope is None:
            query = query.where(ImportJobRow.tenant_ref.is_(None))
        else:
            query = query.where(
                ImportJobRow.tenant_ref == scope["tenant_ref"],
                ImportJobRow.entity_ref == scope["entity_ref"],
                ImportJobRow.store_ref == scope["store_ref"],
            )
        with Session(self.engine) as session:
            existing = session.scalar(query)
            return self._result(existing, duplicate=True) if existing else None

    def preview_file(self, *, filename: str, content: bytes) -> ImportPreview:
        preview, _ = self._inspect_file(filename=filename, content=content)
        return preview

    def _inspect_file(
        self, *, filename: str, content: bytes
    ) -> tuple[ImportPreview, list[dict[str, Any]]]:
        raw_rows = self._read_rows(filename=filename, content=content)
        headers = list(raw_rows[0]) if raw_rows else []
        record_type = detect_record_type(filename, headers)
        detected_mapping = _detect_mapping(headers)
        contract_fields = CONTRACTS[record_type].required_fields | CONTRACTS[record_type].optional_fields
        mapping = {key: value for key, value in detected_mapping.items() if key in contract_fields}
        missing_columns = [
            name for name in sorted(CONTRACTS[record_type].required_fields) if name not in mapping
        ]
        preview = ImportPreview(
            filename=Path(filename).name,
            sha256=hashlib.sha256(content).hexdigest(),
            record_type=record_type.value,
            row_count=len(raw_rows),
            mapping=mapping,
            missing_columns=missing_columns,
            ready=not missing_columns,
        )
        return preview, raw_rows

    def import_file(
        self,
        *,
        filename: str,
        content: bytes,
        evidence_id: str | None = None,
        scope_authority: dict[str, Any] | None = None,
    ) -> ImportResult:
        preview, raw_rows = self._inspect_file(filename=filename, content=content)
        digest = hashlib.sha256(content).hexdigest()
        scope = self._scope_authority(scope_authority)
        if scope is not None and not evidence_id:
            raise ValueError("Native Ozon import requires source Evidence")
        existing = self.find_by_content(
            content,
            scope_authority=scope,
        )
        if existing:
            if scope is not None and (
                existing.scope is None
                or existing.scope["scope_grant_authority_sha256"]
                != scope["scope_grant_authority_sha256"]
                or existing.scope["source_evidence_sha256"]
                != scope["source_evidence_sha256"]
            ):
                raise ValueError(
                    "Ozon import scope or Evidence authority changed"
                )
            return existing

        record_type = OzonRecordType(preview.record_type)
        mapping = preview.mapping
        missing_columns = preview.missing_columns
        file_errors = [{"type": "missing_column", "field": name} for name in missing_columns]
        job_id = new_id("imp")
        accepted = 0
        rejected = 0
        data_rows: list[ImportDataRow] = []
        for row_number, raw in enumerate(raw_rows, start=2):
            normalized, errors = _normalize(
                raw,
                mapping,
                record_type,
                source_row_id=f"report-row:{digest[:16]}:{row_number}",
            )
            accepted += not errors
            rejected += bool(errors)
            data_rows.append(
                ImportDataRow(
                    id=new_id("row"),
                    import_id=job_id,
                    row_number=row_number,
                    record_type=record_type.value,
                    external_id=normalized.get("external_id"),
                    payload_json={str(key): _json_value(value) for key, value in raw.items()},
                    normalized_json=normalized,
                    errors_json=errors,
                )
            )
        status = "rejected" if missing_columns else ("completed_with_errors" if rejected else "completed")
        job = ImportJobRow(
            id=job_id,
            source="ozon-export",
            record_type=record_type.value,
            filename=Path(filename).name,
            sha256=digest,
            status=status,
            row_count=len(raw_rows),
            accepted_count=accepted,
            rejected_count=rejected,
            mapping_json=mapping,
            errors_json=file_errors,
            evidence_id=evidence_id,
            created_at=datetime.now(UTC),
            tenant_ref=(
                scope["tenant_ref"] if scope is not None else None
            ),
            entity_ref=(
                scope["entity_ref"] if scope is not None else None
            ),
            store_ref=(
                scope["store_ref"] if scope is not None else None
            ),
            scope_grant_authority_sha256=(
                scope["scope_grant_authority_sha256"]
                if scope is not None
                else None
            ),
            source_evidence_sha256=(
                scope["source_evidence_sha256"]
                if scope is not None
                else None
            ),
            scope_as_of=(
                self._datetime(scope["scope_as_of"], "scope_as_of")
                if scope is not None
                else None
            ),
        )
        with Session(self.engine, expire_on_commit=False) as session, session.begin():
            session.add(job)
            session.flush()
            session.add_all(data_rows)
        return self._result(job)

    @staticmethod
    def _read_rows(*, filename: str, content: bytes) -> list[dict[str, Any]]:
        if not content:
            raise ValueError("Import file is empty")
        if len(content) > MAX_IMPORT_BYTES:
            raise ValueError("Import file exceeds 20 MB")
        extension = Path(filename).suffix.lower()
        if extension not in {".csv", ".xlsx"}:
            raise ValueError("Only CSV and XLSX imports are supported")
        rows = list(_csv_rows(content) if extension == ".csv" else _xlsx_rows(content))
        if len(rows) > MAX_IMPORT_ROWS:
            raise ValueError("Import file exceeds 50,000 rows")
        return rows

    def get(self, import_id: str) -> ImportResult:
        with Session(self.engine) as session:
            row = session.get(ImportJobRow, import_id)
            if row is None:
                raise KeyError(f"Unknown import: {import_id}")
            return self._result(row)

    @staticmethod
    def _result(row: ImportJobRow, *, duplicate: bool = False) -> ImportResult:
        return ImportResult(
            row.id,
            row.source,
            row.record_type,
            row.filename,
            row.sha256,
            row.status,
            row.row_count,
            row.accepted_count,
            row.rejected_count,
            row.mapping_json,
            row.errors_json,
            row.evidence_id,
            row.created_at.isoformat(),
            duplicate,
            (
                {
                    "tenant_ref": row.tenant_ref,
                    "entity_ref": row.entity_ref,
                    "store_ref": row.store_ref,
                    "scope_grant_authority_sha256": (
                        row.scope_grant_authority_sha256
                    ),
                    "source_evidence_sha256": (
                        row.source_evidence_sha256
                    ),
                    "as_of": OzonImportService._iso(row.scope_as_of),
                    "authority": "native",
                }
                if row.tenant_ref is not None
                else {
                    "tenant_ref": None,
                    "entity_ref": None,
                    "store_ref": None,
                    "scope_grant_authority_sha256": None,
                    "source_evidence_sha256": None,
                    "as_of": None,
                    "authority": "legacy",
                }
            ),
            False,
            False,
        )

    @staticmethod
    def _iso(value: datetime | None) -> str | None:
        if value is None:
            return None
        normalized = (
            value.replace(tzinfo=UTC)
            if value.tzinfo is None
            else value.astimezone(UTC)
        )
        return normalized.isoformat()

    @classmethod
    def _scope_authority(
        cls,
        value: dict[str, Any] | None,
    ) -> dict[str, str] | None:
        if value is None:
            return None
        required = (
            "tenant_ref",
            "entity_ref",
            "store_ref",
            "scope_grant_authority_sha256",
            "source_evidence_sha256",
            "scope_as_of",
        )
        scope = {
            field: str(value.get(field, "")).strip()
            for field in required
        }
        if any(not item for item in scope.values()):
            raise ValueError("Native Ozon import scope is incomplete")
        for field in (
            "scope_grant_authority_sha256",
            "source_evidence_sha256",
        ):
            digest = scope[field].lower()
            if len(digest) != 64 or any(
                character not in "0123456789abcdef"
                for character in digest
            ):
                raise ValueError(f"{field} must be SHA-256")
            scope[field] = digest
        scope["scope_as_of"] = cls._datetime(
            scope["scope_as_of"],
            "scope_as_of",
        ).isoformat()
        return scope

    @staticmethod
    def _datetime(value: str, name: str) -> datetime:
        try:
            parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
        except (AttributeError, ValueError) as exc:
            raise ValueError(f"{name} must be ISO-8601") from exc
        if parsed.tzinfo is None:
            raise ValueError(f"{name} must include timezone")
        return parsed.astimezone(UTC)

from __future__ import annotations

import base64
import binascii
import csv
import hashlib
import io
import json
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .evidence import EvidenceGrade, parse_timestamp
from .evidence_scope import BINDING_CONTRACT
from .security import Principal


class ScopedSellerErpBridge:
    """Admit authorized Seller ERP snapshots and diff them against KJDS truth."""

    CONTRACT_ID = "kjds-scoped-seller-erp-bridge-v1"
    ARTIFACT_CONTRACT_ID = "kjds-seller-erp-bridge-agent-artifact-v1"
    SOURCE_CONTRACT_ID = "kjds-seller-erp-bridge-source-v1"
    REVIEW_CONTRACT_ID = "kjds-seller-erp-bridge-review-v1"
    BINDING_WORKFLOW_CONTRACT_ID = (
        "kjds-seller-erp-bridge-binding-workflow-v1"
    )
    REVOCATION_CONTRACT_ID = "kjds-seller-erp-bridge-revocation-v1"

    SOURCE_NAME = "seller_erp_bridge_source"
    REVIEW_SOURCE_NAME = "seller_erp_bridge_review"
    BINDING_SOURCE_NAME = "seller_erp_bridge_binding"
    REVOCATION_SOURCE_NAME = "seller_erp_bridge_revocation"

    SOURCE_KINDS = frozenset(
        {
            "platform_official_export",
            "seller_erp_formal_export",
            "authorized_adapter_snapshot",
        }
    )
    AUTHORIZATION_MODES = frozenset(
        {
            "first_party_account_export",
            "account_owner_export",
            "public_api_oauth",
            "contracted_api",
            "written_authorization",
        }
    )
    DOMAINS = frozenset({"catalog", "orders", "inventory"})
    MAX_ROWS = 5000
    MAX_COLUMNS = 64
    MAX_CELL_CHARS = 5000
    MAX_AUTHORITY_RECORDS = 2000

    SCHEMAS: dict[str, dict[str, Any]] = {
        "catalog": {
            "version": "seller-erp-bridge-catalog-v1",
            "required": {
                "seller_sku",
                "offer_id",
                "title",
                "status",
            },
            "optional": {"marketplace_sku"},
            "compare": (
                "seller_sku",
                "offer_id",
                "marketplace_sku",
                "title",
                "status",
            ),
        },
        "orders": {
            "version": "seller-erp-bridge-orders-v1",
            "required": {
                "order_external_id",
                "seller_sku",
                "status",
                "quantity",
                "gross_revenue",
                "currency",
                "updated_at",
            },
            "optional": set(),
            "compare": (
                "order_external_id",
                "seller_sku",
                "status",
                "quantity",
                "gross_revenue",
                "currency",
            ),
        },
        "inventory": {
            "version": "seller-erp-bridge-inventory-v1",
            "required": {
                "seller_sku",
                "warehouse_ref",
                "fulfillment_mode",
                "available_quantity",
                "reserved_quantity",
                "in_transit_quantity",
                "damaged_quantity",
                "quarantine_quantity",
                "updated_at",
            },
            "optional": set(),
            "compare": (
                "seller_sku",
                "warehouse_ref",
                "fulfillment_mode",
                "available_quantity",
                "reserved_quantity",
                "in_transit_quantity",
                "damaged_quantity",
                "quarantine_quantity",
            ),
        },
    }
    SECRET_HEADERS = frozenset(
        {
            "access_token",
            "authorization",
            "cookie",
            "cookies",
            "password",
            "refresh_token",
            "secret",
            "session",
            "session_id",
            "token",
        }
    )

    def __init__(
        self,
        *,
        evidence,
        scoped_evidence,
        pim,
        oms,
        inventory,
    ) -> None:
        self.evidence = evidence
        self.scoped_evidence = scoped_evidence
        self.pim = pim
        self.oms = oms
        self.inventory = inventory

    def submit_source(
        self,
        *,
        principal: Principal,
        entity_scope: dict[str, Any],
        store_ref: str,
        provider: str,
        source_kind: str,
        domain: str,
        schema_version: str,
        column_map: dict[str, str],
        exported_at: str,
        authorization_mode: str,
        authorization_evidence_id: str | None,
        effective_until: str | None,
        idempotency_key: str,
        content: bytes,
        filename: str,
        content_type: str,
        worksheet: str | None = None,
    ) -> dict[str, Any]:
        context = self._write_context(
            principal=principal,
            entity_scope=entity_scope,
            store_ref=store_ref,
        )
        provider = self._required(provider, "provider", 160)
        source_kind = self._choice(
            source_kind, "source_kind", self.SOURCE_KINDS
        )
        domain = self._choice(domain, "domain", self.DOMAINS)
        authorization_mode = self._choice(
            authorization_mode,
            "authorization_mode",
            self.AUTHORIZATION_MODES,
        )
        idempotency_key = self._required(
            idempotency_key, "idempotency_key", 300
        )
        exported = parse_timestamp(exported_at, "exported_at")
        if exported > datetime.now(UTC):
            raise ValueError("exported_at cannot be in the future")
        effective_end = (
            parse_timestamp(effective_until, "effective_until")
            if effective_until
            else None
        )
        if effective_end is not None and effective_end <= exported:
            raise ValueError("effective_until must be later than exported_at")
        authorization_evidence_id = (
            self._required(
                authorization_evidence_id,
                "authorization_evidence_id",
                160,
            )
            if authorization_evidence_id
            else None
        )
        if (
            source_kind == "authorized_adapter_snapshot"
            and authorization_evidence_id is None
        ):
            raise ValueError(
                "authorized_adapter_snapshot requires authorization_evidence_id"
            )
        if authorization_evidence_id is not None:
            self.evidence.require_current(
                [authorization_evidence_id],
                as_of=exported,
            )
        normalized_map = self._column_map(
            domain=domain,
            schema_version=schema_version,
            column_map=column_map,
        )
        parsed = self._parse_snapshot(
            content=content,
            filename=filename,
            content_type=content_type,
            domain=domain,
            column_map=normalized_map,
            worksheet=worksheet,
        )
        metadata = {
            "seller_erp_bridge_source_contract_id": self.SOURCE_CONTRACT_ID,
            **context["scope"],
            "scope_grant_authority_sha256": context["authority_sha256"],
            "provider": provider,
            "source_kind": source_kind,
            "domain": domain,
            "schema_version": self.SCHEMAS[domain]["version"],
            "column_map": normalized_map,
            "exported_at": exported.isoformat(),
            "authorization_mode": authorization_mode,
            "authorization_evidence_id": authorization_evidence_id,
            "worksheet": parsed["worksheet"],
            "header_sha256": parsed["header_sha256"],
            "normalized_rows_sha256": parsed["rows_sha256"],
            "row_count": len(parsed["rows"]),
            "retention_class": "operational",
            "legal_hold": False,
            "private_interface_used": False,
            "session_material_stored": False,
        }
        source_ref = (
            f"seller-erp-bridge-source://{principal.tenant_ref}/"
            f"{principal.actor_id}/{idempotency_key}"
        )
        existing = self.evidence.find_by_source_ref(
            source=self.SOURCE_NAME,
            source_ref=source_ref,
        )
        if existing is not None:
            self._require_replay(
                existing=existing,
                content=content,
                effective_at=exported,
                metadata=metadata,
                created_by=principal.actor_id,
            )
            return self._source_projection(
                existing,
                status="pending_independent_review",
                idempotent=True,
            )
        source = self.evidence.capture(
            content=content,
            filename=filename,
            content_type=content_type,
            source=self.SOURCE_NAME,
            source_ref=source_ref,
            grade=EvidenceGrade.B,
            effective_at=exported.isoformat(),
            effective_until=(
                effective_end.isoformat() if effective_end else None
            ),
            created_by=principal.actor_id,
            metadata=metadata,
        )
        self._require_replay(
            existing=source,
            content=content,
            effective_at=exported,
            metadata=metadata,
            created_by=principal.actor_id,
        )
        return self._source_projection(
            source,
            status="pending_independent_review",
            idempotent=False,
        )

    def review_source(
        self,
        *,
        principal: Principal,
        entity_scope: dict[str, Any],
        store_ref: str,
        source_evidence_id: str,
        accepted: bool,
        authentic_original: bool,
        authorization_verified: bool,
        export_scope_matches: bool,
        schema_mapping_verified: bool,
        no_session_or_secret_material: bool,
        rationale: str,
        effective_at: str,
        idempotency_key: str,
    ) -> dict[str, Any]:
        context = self._write_context(
            principal=principal,
            entity_scope=entity_scope,
            store_ref=store_ref,
        )
        source_evidence_id = self._required(
            source_evidence_id, "source_evidence_id", 160
        )
        rationale = self._required(rationale, "rationale", 5000)
        idempotency_key = self._required(
            idempotency_key, "idempotency_key", 300
        )
        reviewed_at = parse_timestamp(effective_at, "effective_at")
        if reviewed_at > datetime.now(UTC):
            raise ValueError("effective_at cannot be in the future")
        self.evidence.require_current(
            [source_evidence_id], as_of=reviewed_at
        )
        source = self.evidence.get(source_evidence_id)
        self._validate_source_record(
            source=source,
            context=context,
            as_of=reviewed_at,
        )
        if reviewed_at < self._stored_time(source.recorded_at):
            raise ValueError(
                "effective_at cannot predate source Evidence recording"
            )
        if principal.actor_id == source.created_by:
            raise PermissionError(
                "Seller ERP source review requires an independent reviewer"
            )
        checks = {
            "authentic_original": authentic_original,
            "authorization_verified": authorization_verified,
            "export_scope_matches": export_scope_matches,
            "schema_mapping_verified": schema_mapping_verified,
            "no_session_or_secret_material": no_session_or_secret_material,
        }
        if accepted and not all(checks.values()):
            raise ValueError(
                "Accepted Seller ERP source review requires every check to pass"
            )
        payload = {
            "seller_erp_bridge_review_contract_id": self.REVIEW_CONTRACT_ID,
            **context["scope"],
            "source_evidence_id": source.id,
            "source_evidence_sha256": source.sha256,
            "decision": "accepted" if accepted else "rejected",
            "reviewed_by": principal.actor_id,
            "checks": checks,
            "rationale": rationale,
        }
        content = self._json_bytes(payload)
        source_ref = (
            f"seller-erp-bridge-review://{principal.tenant_ref}/"
            f"{principal.actor_id}/{idempotency_key}"
        )
        existing = self.evidence.find_by_source_ref(
            source=self.REVIEW_SOURCE_NAME,
            source_ref=source_ref,
        )
        if existing is not None:
            self._require_replay(
                existing=existing,
                content=content,
                effective_at=reviewed_at,
                metadata={
                    **payload,
                    "retention_class": "compliance",
                    "legal_hold": False,
                },
                created_by=principal.actor_id,
            )
            return self._review_projection(
                source=source,
                review=existing,
                idempotent=True,
            )
        review = self.evidence.capture(
            content=content,
            filename=f"{source.id}-seller-erp-review.json",
            content_type="application/json",
            source=self.REVIEW_SOURCE_NAME,
            source_ref=source_ref,
            grade=EvidenceGrade.A,
            effective_at=reviewed_at.isoformat(),
            effective_until=None,
            created_by=principal.actor_id,
            metadata={
                **payload,
                "retention_class": "compliance",
                "legal_hold": False,
            },
        )
        self.evidence.link(
            evidence_id=review.id,
            target_type="evidence",
            target_id=source.id,
            relationship="seller_erp_bridge_review",
            created_by=principal.actor_id,
        )
        return self._review_projection(
            source=source,
            review=review,
            idempotent=False,
        )

    def bind_source(
        self,
        *,
        principal: Principal,
        entity_scope: dict[str, Any],
        store_ref: str,
        source_evidence_id: str,
        review_evidence_id: str,
        effective_at: str,
        effective_until: str | None,
        idempotency_key: str,
    ) -> dict[str, Any]:
        context = self._write_context(
            principal=principal,
            entity_scope=entity_scope,
            store_ref=store_ref,
        )
        source_evidence_id = self._required(
            source_evidence_id, "source_evidence_id", 160
        )
        review_evidence_id = self._required(
            review_evidence_id, "review_evidence_id", 160
        )
        idempotency_key = self._required(
            idempotency_key, "idempotency_key", 300
        )
        bound_at = parse_timestamp(effective_at, "effective_at")
        if bound_at > datetime.now(UTC):
            raise ValueError("effective_at cannot be in the future")
        bound_until = (
            parse_timestamp(effective_until, "effective_until")
            if effective_until
            else None
        )
        if bound_until is not None and bound_until <= bound_at:
            raise ValueError("effective_until must be later than effective_at")
        self.evidence.require_current(
            [source_evidence_id, review_evidence_id],
            as_of=bound_at,
        )
        source = self.evidence.get(source_evidence_id)
        review = self.evidence.get(review_evidence_id)
        self._validate_source_record(
            source=source,
            context=context,
            as_of=bound_at,
        )
        self._validate_review_record(
            review=review,
            source=source,
            context=context,
            accepted_required=True,
        )
        if bound_at < self._stored_time(review.recorded_at):
            raise ValueError(
                "effective_at cannot predate review Evidence recording"
            )
        if principal.actor_id in {
            source.created_by,
            review.created_by,
        }:
            raise PermissionError(
                "Seller ERP binding recorder must be independent from "
                "source submitter and reviewer"
            )
        payload = {
            "evidence_scope_contract_id": BINDING_CONTRACT,
            "seller_erp_bridge_binding_contract_id": (
                self.BINDING_WORKFLOW_CONTRACT_ID
            ),
            **context["scope"],
            "target_evidence_id": source.id,
            "target_evidence_sha256": source.sha256,
            "review_evidence_id": review.id,
            "review_evidence_sha256": review.sha256,
            "reviewed_by": review.created_by,
            "recorded_by": principal.actor_id,
        }
        content = self._json_bytes(payload)
        source_ref = (
            f"seller-erp-bridge-binding://{principal.tenant_ref}/"
            f"{principal.actor_id}/{idempotency_key}"
        )
        existing = self.evidence.find_by_source_ref(
            source=self.BINDING_SOURCE_NAME,
            source_ref=source_ref,
        )
        metadata = {
            **payload,
            "retention_class": "compliance",
            "legal_hold": False,
        }
        if existing is not None:
            self._require_replay(
                existing=existing,
                content=content,
                effective_at=bound_at,
                metadata=metadata,
                created_by=principal.actor_id,
            )
            return self._binding_projection(
                source=source,
                review=review,
                binding=existing,
                idempotent=True,
            )
        binding = self.evidence.capture(
            content=content,
            filename=f"{source.id}-seller-erp-binding.json",
            content_type="application/json",
            source=self.BINDING_SOURCE_NAME,
            source_ref=source_ref,
            grade=EvidenceGrade.A,
            effective_at=bound_at.isoformat(),
            effective_until=(
                bound_until.isoformat() if bound_until else None
            ),
            created_by=principal.actor_id,
            metadata=metadata,
        )
        self.evidence.link(
            evidence_id=binding.id,
            target_type="evidence",
            target_id=source.id,
            relationship="seller_erp_bridge_scope_binding",
            created_by=principal.actor_id,
        )
        self.evidence.link(
            evidence_id=binding.id,
            target_type="evidence",
            target_id=review.id,
            relationship="seller_erp_bridge_binding_review",
            created_by=principal.actor_id,
        )
        return self._binding_projection(
            source=source,
            review=review,
            binding=binding,
            idempotent=False,
        )

    def revoke_source(
        self,
        *,
        principal: Principal,
        entity_scope: dict[str, Any],
        store_ref: str,
        source_evidence_id: str,
        reason: str,
        effective_at: str,
        idempotency_key: str,
    ) -> dict[str, Any]:
        context = self._write_context(
            principal=principal,
            entity_scope=entity_scope,
            store_ref=store_ref,
        )
        source_evidence_id = self._required(
            source_evidence_id, "source_evidence_id", 160
        )
        reason = self._required(reason, "reason", 2000)
        idempotency_key = self._required(
            idempotency_key, "idempotency_key", 300
        )
        revoked_at = parse_timestamp(effective_at, "effective_at")
        if revoked_at > datetime.now(UTC):
            raise ValueError("effective_at cannot be in the future")
        self.evidence.require_current(
            [source_evidence_id], as_of=revoked_at
        )
        source = self.evidence.get(source_evidence_id)
        self._validate_source_record(
            source=source,
            context=context,
            as_of=revoked_at,
        )
        if revoked_at < self._stored_time(source.recorded_at):
            raise ValueError(
                "effective_at cannot predate source Evidence recording"
            )
        if principal.actor_id == source.created_by:
            raise PermissionError(
                "Seller ERP source revocation requires independent compliance"
            )
        payload = {
            "seller_erp_bridge_revocation_contract_id": (
                self.REVOCATION_CONTRACT_ID
            ),
            **context["scope"],
            "target_evidence_id": source.id,
            "target_evidence_sha256": source.sha256,
            "reason": reason,
            "revoked_by": principal.actor_id,
        }
        content = self._json_bytes(payload)
        source_ref = (
            f"seller-erp-bridge-revocation://{principal.tenant_ref}/"
            f"{principal.actor_id}/{idempotency_key}"
        )
        metadata = {
            **payload,
            "retention_class": "compliance",
            "legal_hold": False,
        }
        existing = self.evidence.find_by_source_ref(
            source=self.REVOCATION_SOURCE_NAME,
            source_ref=source_ref,
        )
        if existing is not None:
            self._require_replay(
                existing=existing,
                content=content,
                effective_at=revoked_at,
                metadata=metadata,
                created_by=principal.actor_id,
            )
            return self._revocation_projection(
                source=source,
                revocation=existing,
                idempotent=True,
            )
        revocation = self.evidence.capture(
            content=content,
            filename=f"{source.id}-seller-erp-revocation.json",
            content_type="application/json",
            source=self.REVOCATION_SOURCE_NAME,
            source_ref=source_ref,
            grade=EvidenceGrade.A,
            effective_at=revoked_at.isoformat(),
            effective_until=None,
            created_by=principal.actor_id,
            metadata=metadata,
        )
        self.evidence.link(
            evidence_id=revocation.id,
            target_type="evidence",
            target_id=source.id,
            relationship="seller_erp_bridge_revocation",
            created_by=principal.actor_id,
        )
        return self._revocation_projection(
            source=source,
            revocation=revocation,
            idempotent=False,
        )

    def reconcile(
        self,
        *,
        principal: Principal,
        entity_scope: dict[str, Any],
        store_ref: str,
        as_of: datetime,
        source_evidence_id: str | None = None,
        page_size: int = 100,
        cursor: str | None = None,
        query: str | None = None,
        state: str | None = None,
    ) -> dict[str, Any]:
        if not 1 <= page_size <= 500:
            raise ValueError("page_size must be between 1 and 500")
        if state not in {
            None,
            "matched",
            "source_only",
            "canonical_only",
            "conflict",
            "blocked",
        }:
            raise ValueError("Unsupported reconciliation state")
        context = self._read_context(
            principal=principal,
            entity_scope=entity_scope,
            store_ref=store_ref,
            as_of=as_of,
        )
        source_id = str(source_evidence_id or "").strip() or None
        normalized_cursor = str(cursor or "").strip() or None
        normalized_query = str(query or "").strip().casefold()
        if context["status"] != "ready":
            return self._empty_result(
                context=context,
                source_evidence_id=source_id,
                page_size=page_size,
                cursor=normalized_cursor,
                query=normalized_query,
                state=state,
            )
        if source_id is None:
            return self._empty_result(
                context=context,
                source_evidence_id=None,
                page_size=page_size,
                cursor=normalized_cursor,
                query=normalized_query,
                state=state,
                gap="seller_erp_bridge_source_missing",
            )

        admission = self._admit_source(
            principal=principal,
            entity_scope=entity_scope,
            store_ref=store_ref,
            as_of=context["cutoff"],
            source_evidence_id=source_id,
            context=context,
        )
        if admission["status"] != "ready":
            return self._blocked_result(
                context=context,
                source_evidence_id=source_id,
                page_size=page_size,
                cursor=normalized_cursor,
                query=normalized_query,
                state=state,
                gaps=admission["source_gaps"],
                authority=admission["authority"],
            )
        source = admission["source"]
        metadata = source.metadata
        try:
            content, content_record = self.evidence.content(source.id)
            if (
                content_record.sha256 != source.sha256
                or self._sha(content) != source.sha256
            ):
                raise ValueError("source Evidence content hash mismatch")
            parsed = self._parse_snapshot(
                content=content,
                filename=source.filename,
                content_type=source.content_type,
                domain=metadata["domain"],
                column_map=metadata["column_map"],
                worksheet=metadata.get("worksheet"),
            )
            if (
                parsed["header_sha256"] != metadata.get("header_sha256")
                or parsed["rows_sha256"]
                != metadata.get("normalized_rows_sha256")
                or len(parsed["rows"]) != metadata.get("row_count")
            ):
                raise ValueError("source normalized snapshot drift")
        except (KeyError, RuntimeError, TypeError, ValueError) as exc:
            return self._blocked_result(
                context=context,
                source_evidence_id=source_id,
                page_size=page_size,
                cursor=normalized_cursor,
                query=normalized_query,
                state=state,
                gaps=[
                    self._safe_gap(
                        "seller_erp_bridge_source_parse_or_schema_invalid",
                        exc,
                    )
                ],
                authority=admission["authority"],
            )

        domain = str(metadata["domain"])
        upstream = self._upstream(
            domain=domain,
            principal=principal,
            entity_scope=entity_scope,
            store_ref=store_ref,
            as_of=context["cutoff"],
        )
        upstream_gaps = self._validate_upstream(
            domain=domain,
            upstream=upstream,
            context=context,
        )
        if upstream_gaps:
            return self._blocked_result(
                context=context,
                source_evidence_id=source_id,
                page_size=page_size,
                cursor=normalized_cursor,
                query=normalized_query,
                state=state,
                gaps=upstream_gaps,
                authority=admission["authority"],
                upstream=upstream,
            )
        try:
            canonical_rows = self._canonical_rows(
                domain=domain,
                upstream=upstream,
            )
            items = self._diff(
                domain=domain,
                source_rows=parsed["rows"],
                canonical_rows=canonical_rows,
            )
        except ValueError as exc:
            return self._blocked_result(
                context=context,
                source_evidence_id=source_id,
                page_size=page_size,
                cursor=normalized_cursor,
                query=normalized_query,
                state=state,
                gaps=[
                    self._safe_gap(
                        "seller_erp_bridge_canonical_projection_invalid",
                        exc,
                    )
                ],
                authority=admission["authority"],
                upstream=upstream,
            )

        if normalized_query:
            items = [
                item
                for item in items
                if normalized_query
                in json.dumps(
                    {
                        "key": item["canonical_key"],
                        "source": item["source"],
                        "canonical": item["canonical"],
                    },
                    ensure_ascii=False,
                    sort_keys=True,
                ).casefold()
            ]
        if state:
            items = [item for item in items if item["state"] == state]
        total_items = len(items)
        state_counts = {
            name: sum(item["state"] == name for item in items)
            for name in (
                "matched",
                "source_only",
                "canonical_only",
                "conflict",
                "blocked",
            )
        }
        if normalized_cursor:
            cursor_key = self._decode_cursor(normalized_cursor)
            items = [
                item
                for item in items
                if (item["state"], item["canonical_key"]) > cursor_key
            ]
        page = items[:page_size]
        next_cursor = (
            self._encode_cursor(
                (page[-1]["state"], page[-1]["canonical_key"])
            )
            if len(items) > page_size and page
            else None
        )
        source_gaps = []
        if upstream["status"] == "no_data" and parsed["rows"]:
            source_gaps.append(
                f"canonical_{domain}_authority_no_data"
            )
        if state_counts["source_only"]:
            source_gaps.append("seller_erp_source_only_rows_require_review")
        if state_counts["canonical_only"]:
            source_gaps.append(
                "canonical_only_rows_missing_from_seller_erp_snapshot"
            )
        if state_counts["conflict"]:
            source_gaps.append("seller_erp_canonical_conflicts_require_review")
        status = (
            "no_data"
            if total_items == 0
            else "partial"
            if source_gaps
            else "ready"
        )
        core = {
            "contract_id": self.CONTRACT_ID,
            "status": status,
            "as_of": context["cutoff"].isoformat(),
            "scope": context["scope"],
            "source": {
                "evidence_id": source.id,
                "sha256": source.sha256,
                "provider": metadata["provider"],
                "source_kind": metadata["source_kind"],
                "domain": domain,
                "schema_version": metadata["schema_version"],
                "exported_at": metadata["exported_at"],
                "authorization_mode": metadata["authorization_mode"],
                "row_count": metadata["row_count"],
            },
            "authority": admission["authority"],
            "query": {
                "page_size": page_size,
                "cursor": normalized_cursor,
                "next_cursor": next_cursor,
                "search": normalized_query or None,
                "state": state,
            },
            "counts": {
                "total_diff_items": total_items,
                "page_diff_items": len(page),
                "source_rows": len(parsed["rows"]),
                "canonical_rows": len(canonical_rows),
                **state_counts,
            },
            "diff_items": page,
            "source_gaps": sorted(source_gaps),
            "blockers": [
                self._blocker(code) for code in sorted(source_gaps)
            ],
            "upstream_authority": {
                "contract_id": upstream.get("contract_id"),
                "status": upstream.get("status"),
                "snapshot_sha256": upstream.get("snapshot_sha256"),
            },
            "control_envelope": self._control(input_read=True),
        }
        input_hash = self._hash(core)
        core["agent_artifact"] = {
            "contract_id": self.ARTIFACT_CONTRACT_ID,
            "input_snapshot_sha256": input_hash,
            "artifact_sha256": self._hash(
                {
                    "contract_id": self.ARTIFACT_CONTRACT_ID,
                    "input_snapshot_sha256": input_hash,
                    "suggestions": [
                        {
                            "canonical_key": item["canonical_key"],
                            "state": item["state"],
                            "owner": item["owner"],
                            "next": item["next"],
                        }
                        for item in page
                        if item["state"] != "matched"
                    ],
                }
            ),
            "authority": "decision_support_and_internal_task_suggestion_only",
            "self_approval_allowed": False,
            "permit_issue_allowed": False,
            "formal_fact_promotion_allowed": False,
            "external_write_allowed": False,
        }
        core["snapshot_sha256"] = self._hash(core)
        return core

    def _admit_source(
        self,
        *,
        principal: Principal,
        entity_scope: dict[str, Any],
        store_ref: str,
        as_of: datetime,
        source_evidence_id: str,
        context: dict[str, Any],
    ) -> dict[str, Any]:
        gaps: list[str] = []
        authority: dict[str, Any] = {
            "source_evidence_id": source_evidence_id,
            "source_evidence_sha256": None,
            "review_evidence_id": None,
            "binding_evidence_id": None,
            "revocation_evidence_id": None,
            "three_party_independence": False,
        }
        try:
            self.evidence.require_current(
                [source_evidence_id], as_of=as_of
            )
            source = self.evidence.get(source_evidence_id)
            self._validate_source_record(
                source=source,
                context=context,
                as_of=as_of,
            )
            if self._stored_time(source.recorded_at) > as_of:
                raise ValueError("source Evidence recorded after as_of")
            authority["source_evidence_sha256"] = source.sha256
        except (KeyError, RuntimeError, TypeError, ValueError) as exc:
            gaps.append(
                self._safe_gap(
                    "seller_erp_bridge_source_evidence_invalid", exc
                )
            )
            return {
                "status": "blocked",
                "source": None,
                "source_gaps": sorted(set(gaps)),
                "authority": authority,
            }

        review_records = self._target_records(
            source_name=self.REVIEW_SOURCE_NAME,
            target_id=source.id,
            as_of=as_of,
        )
        if review_records["truncated"]:
            gaps.append("seller_erp_bridge_review_scan_truncated")
        review = review_records["latest"]
        if review is None:
            gaps.append("seller_erp_bridge_independent_review_missing")
        else:
            try:
                self.evidence.require_current([review.id], as_of=as_of)
                self._validate_review_record(
                    review=review,
                    source=source,
                    context=context,
                    accepted_required=True,
                )
                authority["review_evidence_id"] = review.id
            except (KeyError, RuntimeError, TypeError, ValueError) as exc:
                gaps.append(
                    self._safe_gap(
                        "seller_erp_bridge_latest_review_invalid", exc
                    )
                )

        revocations = self._target_records(
            source_name=self.REVOCATION_SOURCE_NAME,
            target_id=source.id,
            as_of=as_of,
        )
        if revocations["truncated"]:
            gaps.append("seller_erp_bridge_revocation_scan_truncated")
        revocation = revocations["latest"]
        if revocation is not None:
            authority["revocation_evidence_id"] = revocation.id
            try:
                self.evidence.require_current(
                    [revocation.id], as_of=as_of
                )
                self._validate_revocation_record(
                    revocation=revocation,
                    source=source,
                    context=context,
                )
                gaps.append("seller_erp_bridge_source_revoked")
            except (KeyError, RuntimeError, TypeError, ValueError) as exc:
                gaps.append(
                    self._safe_gap(
                        "seller_erp_bridge_revocation_evidence_invalid",
                        exc,
                    )
                )

        binding = None
        if review is not None and not any(
            "review" in gap for gap in gaps
        ):
            binding = self._dedicated_binding(
                source=source,
                review=review,
                as_of=as_of,
            )
        if binding is None:
            gaps.append("seller_erp_bridge_scope_binding_missing")
        else:
            try:
                scoped = self.scoped_evidence.project(
                    evidence_ids=[source.id, binding.id],
                    principal=principal,
                    entity_scope=entity_scope,
                    store_ref=store_ref,
                    as_of=as_of,
                )
                if scoped.get("status") != "ready":
                    gaps.extend(
                        scoped.get("source_gaps")
                        or ["seller_erp_bridge_scope_binding_invalid"]
                    )
                else:
                    self._validate_binding_record(
                        binding=binding,
                        source=source,
                        review=review,
                        context=context,
                    )
                    authority["binding_evidence_id"] = binding.id
                    authority["three_party_independence"] = len(
                        {
                            source.created_by,
                            review.created_by,
                            binding.created_by,
                        }
                    ) == 3
                    if not authority["three_party_independence"]:
                        gaps.append(
                            "seller_erp_bridge_three_party_independence_missing"
                        )
            except (KeyError, RuntimeError, TypeError, ValueError) as exc:
                gaps.append(
                    self._safe_gap(
                        "seller_erp_bridge_scope_binding_invalid", exc
                    )
                )
        return {
            "status": "blocked" if gaps else "ready",
            "source": source,
            "source_gaps": sorted(set(gaps)),
            "authority": authority,
        }

    def _dedicated_binding(
        self,
        *,
        source,
        review,
        as_of: datetime,
    ):
        ids = self.evidence.find_binding_ids(
            target_evidence_ids=[source.id],
            binding_contract_id=BINDING_CONTRACT,
            as_of=as_of,
        )
        records = []
        for evidence_id in ids:
            try:
                record = self.evidence.get(evidence_id)
            except (KeyError, RuntimeError):
                continue
            metadata = record.metadata
            if (
                record.source == self.BINDING_SOURCE_NAME
                and metadata.get(
                    "seller_erp_bridge_binding_contract_id"
                )
                == self.BINDING_WORKFLOW_CONTRACT_ID
                and metadata.get("review_evidence_id") == review.id
                and self._stored_time(record.recorded_at) <= as_of
            ):
                records.append(record)
        if not records:
            return None
        return sorted(records, key=self._record_key)[-1]

    def _target_records(
        self,
        *,
        source_name: str,
        target_id: str,
        as_of: datetime,
    ) -> dict[str, Any]:
        records = self.evidence.list_by_source(
            source_name,
            limit=self.MAX_AUTHORITY_RECORDS,
        )
        matching = [
            item
            for item in records
            if (
                item.metadata.get("target_evidence_id") == target_id
                or item.metadata.get("source_evidence_id") == target_id
            )
            and self._stored_time(item.effective_at) <= as_of
            and self._stored_time(item.recorded_at) <= as_of
        ]
        return {
            "latest": (
                sorted(matching, key=self._record_key)[-1]
                if matching
                else None
            ),
            "truncated": len(records) >= self.MAX_AUTHORITY_RECORDS,
        }

    def _validate_source_record(
        self,
        *,
        source,
        context: dict[str, Any],
        as_of: datetime,
    ) -> None:
        metadata = source.metadata
        if source.source != self.SOURCE_NAME:
            raise ValueError("source workflow mismatch")
        if source.grade.value not in {"A", "B"}:
            raise ValueError("source Evidence grade is insufficient")
        if (
            metadata.get("seller_erp_bridge_source_contract_id")
            != self.SOURCE_CONTRACT_ID
        ):
            raise ValueError("source contract mismatch")
        self._require_scope(metadata, context["scope"])
        if (
            metadata.get("scope_grant_authority_sha256")
            != context["authority_sha256"]
        ):
            raise ValueError("source scope authority hash mismatch")
        domain = self._choice(metadata.get("domain"), "domain", self.DOMAINS)
        if metadata.get("schema_version") != self.SCHEMAS[domain]["version"]:
            raise ValueError("source schema version mismatch")
        self._choice(
            metadata.get("source_kind"),
            "source_kind",
            self.SOURCE_KINDS,
        )
        self._choice(
            metadata.get("authorization_mode"),
            "authorization_mode",
            self.AUTHORIZATION_MODES,
        )
        exported = parse_timestamp(
            str(metadata.get("exported_at") or ""),
            "source exported_at",
        )
        if exported > as_of:
            raise ValueError("source exported after as_of")
        if metadata.get("private_interface_used") is not False:
            raise ValueError("private interface source is prohibited")
        if metadata.get("session_material_stored") is not False:
            raise ValueError("session material source is prohibited")
        if (
            metadata.get("source_kind") == "authorized_adapter_snapshot"
            and not metadata.get("authorization_evidence_id")
        ):
            raise ValueError("adapter authorization Evidence is missing")
        self._column_map(
            domain=domain,
            schema_version=str(metadata.get("schema_version") or ""),
            column_map=metadata.get("column_map"),
        )

    def _validate_review_record(
        self,
        *,
        review,
        source,
        context: dict[str, Any],
        accepted_required: bool,
    ) -> None:
        metadata = review.metadata
        if review.source != self.REVIEW_SOURCE_NAME:
            raise ValueError("review workflow mismatch")
        if review.grade.value != EvidenceGrade.A.value:
            raise ValueError("review requires grade A Evidence")
        if (
            metadata.get("seller_erp_bridge_review_contract_id")
            != self.REVIEW_CONTRACT_ID
        ):
            raise ValueError("review contract mismatch")
        self._require_scope(metadata, context["scope"])
        if (
            metadata.get("source_evidence_id") != source.id
            or metadata.get("source_evidence_sha256") != source.sha256
        ):
            raise ValueError("review target mismatch")
        if review.created_by == source.created_by:
            raise ValueError("review independence missing")
        if metadata.get("reviewed_by") != review.created_by:
            raise ValueError("reviewer identity mismatch")
        checks = metadata.get("checks")
        if not isinstance(checks, dict) or set(checks) != {
            "authentic_original",
            "authorization_verified",
            "export_scope_matches",
            "schema_mapping_verified",
            "no_session_or_secret_material",
        }:
            raise ValueError("review checks are incomplete")
        if accepted_required and (
            metadata.get("decision") != "accepted"
            or not all(value is True for value in checks.values())
        ):
            raise ValueError("latest review is not accepted")

    def _validate_binding_record(
        self,
        *,
        binding,
        source,
        review,
        context: dict[str, Any],
    ) -> None:
        metadata = binding.metadata
        if (
            binding.source != self.BINDING_SOURCE_NAME
            or binding.grade.value != EvidenceGrade.A.value
        ):
            raise ValueError("binding workflow or grade mismatch")
        if (
            metadata.get("evidence_scope_contract_id")
            != BINDING_CONTRACT
            or metadata.get("seller_erp_bridge_binding_contract_id")
            != self.BINDING_WORKFLOW_CONTRACT_ID
        ):
            raise ValueError("binding contract mismatch")
        self._require_scope(metadata, context["scope"])
        if (
            metadata.get("target_evidence_id") != source.id
            or metadata.get("target_evidence_sha256") != source.sha256
            or metadata.get("review_evidence_id") != review.id
            or metadata.get("review_evidence_sha256") != review.sha256
            or metadata.get("reviewed_by") != review.created_by
        ):
            raise ValueError("binding target or review mismatch")
        if len(
            {source.created_by, review.created_by, binding.created_by}
        ) != 3:
            raise ValueError("binding three-party independence missing")

    def _validate_revocation_record(
        self,
        *,
        revocation,
        source,
        context: dict[str, Any],
    ) -> None:
        metadata = revocation.metadata
        if (
            revocation.source != self.REVOCATION_SOURCE_NAME
            or revocation.grade.value != EvidenceGrade.A.value
            or metadata.get(
                "seller_erp_bridge_revocation_contract_id"
            )
            != self.REVOCATION_CONTRACT_ID
        ):
            raise ValueError("revocation workflow mismatch")
        self._require_scope(metadata, context["scope"])
        if (
            metadata.get("target_evidence_id") != source.id
            or metadata.get("target_evidence_sha256") != source.sha256
        ):
            raise ValueError("revocation target mismatch")
        if revocation.created_by == source.created_by:
            raise ValueError("revocation independence missing")

    def _upstream(
        self,
        *,
        domain: str,
        principal: Principal,
        entity_scope: dict[str, Any],
        store_ref: str,
        as_of: datetime,
    ) -> dict[str, Any]:
        if domain == "catalog":
            return self.pim.project(
                principal=principal,
                entity_scope=entity_scope,
                store_ref=store_ref,
                as_of=as_of,
                page_size=500,
            )
        if domain == "orders":
            return self.oms.workspace(
                principal=principal,
                entity_scope=entity_scope,
                store_ref=store_ref,
                as_of=as_of,
                page_size=500,
            )
        return self.inventory.workspace(
            principal=principal,
            entity_scope=entity_scope,
            store_ref=store_ref,
            as_of=as_of,
            page_size=500,
        )

    def _validate_upstream(
        self,
        *,
        domain: str,
        upstream: dict[str, Any],
        context: dict[str, Any],
    ) -> list[str]:
        expected = {
            "catalog": "kjds-native-scoped-pim-v1",
            "orders": "kjds-native-scoped-oms-v1",
            "inventory": "kjds-native-scoped-inventory-fulfillment-v1",
        }[domain]
        gaps = []
        if upstream.get("contract_id") != expected:
            gaps.append("seller_erp_bridge_upstream_contract_conflict")
        if upstream.get("status") not in {
            "ready",
            "partial",
            "no_data",
        }:
            gaps.append("seller_erp_bridge_upstream_authority_blocked")
        if upstream.get("as_of") != context["cutoff"].isoformat():
            gaps.append("seller_erp_bridge_upstream_as_of_conflict")
        scope = upstream.get("scope")
        if not isinstance(scope, dict):
            gaps.append("seller_erp_bridge_upstream_scope_missing")
        else:
            for key, value in context["scope"].items():
                if scope.get(key) != value:
                    gaps.append(
                        f"seller_erp_bridge_upstream_{key}_conflict"
                    )
        snapshot = upstream.get("snapshot_sha256")
        if not isinstance(snapshot, str) or len(snapshot) != 64:
            gaps.append("seller_erp_bridge_upstream_snapshot_invalid")
        return sorted(set(gaps))

    def _canonical_rows(
        self, *, domain: str, upstream: dict[str, Any]
    ) -> list[dict[str, Any]]:
        if upstream["status"] == "no_data":
            return []
        if domain == "catalog":
            result = []
            for group in upstream.get("product_groups", []):
                product = group["product"]
                listings = group.get("listings") or [None]
                for listing in listings:
                    row = {
                        "seller_sku": str(product.get("sku") or ""),
                        "offer_id": (
                            str(listing.get("offer_id") or "")
                            if listing
                            else ""
                        ),
                        "marketplace_sku": (
                            str(
                                listing.get("marketplace_sku") or ""
                            )
                            if listing
                            else ""
                        ),
                        "title": str(product.get("name") or ""),
                        "status": str(
                            (
                                listing.get("listing_status")
                                if listing
                                else product.get("status")
                            )
                            or ""
                        ).lower(),
                    }
                    row["canonical_key"] = self._key("catalog", row)
                    result.append(row)
            return self._unique_rows(result)
        if domain == "orders":
            result = []
            for order in upstream.get("orders", []):
                event = order.get("current_event") or {}
                row = {
                    "order_external_id": str(
                        order.get("external_id") or ""
                    ),
                    "seller_sku": str(order.get("sku") or ""),
                    "status": str(
                        order.get("current_state") or ""
                    ).lower(),
                    "quantity": int(event.get("quantity") or 0),
                    "gross_revenue": self._decimal_text(
                        event.get("amount") or "0",
                        "gross_revenue",
                    ),
                    "currency": str(
                        event.get("currency") or ""
                    ).upper(),
                    "updated_at": str(event.get("effective_at") or ""),
                }
                row["canonical_key"] = self._key("orders", row)
                result.append(row)
            return self._unique_rows(result)
        result = []
        for cell in upstream.get("inventory_cells", []):
            current = cell.get("current_snapshot") or {}
            if cell.get("projection_status") == "blocked":
                raise ValueError(
                    "blocked inventory cell cannot be reconciled"
                )
            quantities = current.get("quantities") or {}
            row = {
                "seller_sku": str(current.get("sku") or ""),
                "warehouse_ref": str(
                    current.get("warehouse_ref") or ""
                ),
                "fulfillment_mode": str(
                    current.get("fulfillment_mode") or ""
                ),
                "available_quantity": int(
                    quantities.get("available_quantity") or 0
                ),
                "reserved_quantity": int(
                    quantities.get("reserved_quantity") or 0
                ),
                "in_transit_quantity": int(
                    quantities.get("in_transit_quantity") or 0
                ),
                "damaged_quantity": int(
                    quantities.get("damaged_quantity") or 0
                ),
                "quarantine_quantity": int(
                    quantities.get("quarantine_quantity") or 0
                ),
                "updated_at": str(current.get("effective_at") or ""),
            }
            row["canonical_key"] = self._key("inventory", row)
            result.append(row)
        return self._unique_rows(result)

    def _diff(
        self,
        *,
        domain: str,
        source_rows: list[dict[str, Any]],
        canonical_rows: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        sources = {row["canonical_key"]: row for row in source_rows}
        canonical = {row["canonical_key"]: row for row in canonical_rows}
        items = []
        for key in sorted(set(sources) | set(canonical)):
            source = sources.get(key)
            current = canonical.get(key)
            field_diffs = []
            if source is not None and current is not None:
                for field in self.SCHEMAS[domain]["compare"]:
                    if source.get(field) != current.get(field):
                        field_diffs.append(
                            {
                                "field": field,
                                "source_value": source.get(field),
                                "canonical_value": current.get(field),
                            }
                        )
            state = (
                "source_only"
                if current is None
                else "canonical_only"
                if source is None
                else "conflict"
                if field_diffs
                else "matched"
            )
            owner = (
                "integration-governance"
                if state in {"source_only", "canonical_only"}
                else "domain-data-owner"
                if state == "conflict"
                else "evidence-governance"
            )
            item = {
                "domain": domain,
                "canonical_key": key,
                "state": state,
                "source": source,
                "canonical": current,
                "field_diffs": field_diffs,
                "owner": owner,
                "sla": (
                    "before any formal Fact promotion or operating decision"
                ),
                "next": (
                    "No action; retain the matched immutable comparison."
                    if state == "matched"
                    else "Review identity and source mapping; do not create a Fact."
                    if state in {"source_only", "canonical_only"}
                    else "Resolve every field conflict against primary Evidence."
                ),
                "next_workspace": (
                    "/pim"
                    if domain == "catalog"
                    else "/oms"
                    if domain == "orders"
                    else "/inventory"
                ),
            }
            item["item_sha256"] = self._hash(item)
            items.append(item)
        items.sort(key=lambda item: (item["state"], item["canonical_key"]))
        return items

    def _parse_snapshot(
        self,
        *,
        content: bytes,
        filename: str,
        content_type: str,
        domain: str,
        column_map: dict[str, str],
        worksheet: str | None,
    ) -> dict[str, Any]:
        if not content:
            raise ValueError("Seller ERP source file cannot be empty")
        extension = Path(filename).suffix.lower()
        if extension == ".csv":
            headers, raw_rows, resolved_sheet = self._read_csv(content)
        elif extension == ".xlsx":
            headers, raw_rows, resolved_sheet = self._read_xlsx(
                content, worksheet
            )
        else:
            raise ValueError("Seller ERP source must be .csv or .xlsx")
        if len(headers) > self.MAX_COLUMNS:
            raise ValueError("Seller ERP source has too many columns")
        folded = [header.casefold() for header in headers]
        if len(folded) != len(set(folded)):
            raise ValueError("Seller ERP source headers must be unique")
        secret_headers = sorted(
            set(folded).intersection(self.SECRET_HEADERS)
        )
        if secret_headers:
            raise ValueError(
                "Seller ERP source contains prohibited secret/session columns"
            )
        missing = sorted(
            source_header
            for source_header in column_map.values()
            if source_header not in headers
        )
        if missing:
            raise ValueError(
                "Seller ERP source is missing mapped columns: "
                + ", ".join(missing)
            )
        normalized = []
        for index, raw in enumerate(raw_rows, start=2):
            row = self._normalize_row(
                domain=domain,
                raw=raw,
                column_map=column_map,
                row_number=index,
            )
            normalized.append(row)
        normalized = self._unique_rows(normalized)
        return {
            "worksheet": resolved_sheet,
            "headers": headers,
            "header_sha256": self._hash(headers),
            "rows": normalized,
            "rows_sha256": self._hash(normalized),
            "content_type": content_type,
        }

    def _read_csv(
        self, content: bytes
    ) -> tuple[list[str], list[dict[str, Any]], None]:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ValueError("CSV must use UTF-8 encoding") from exc
        if "\x00" in text:
            raise ValueError("CSV contains NUL bytes")
        sample = text[:8192]
        try:
            dialect = csv.Sniffer().sniff(
                sample, delimiters=",;\t"
            )
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        if not reader.fieldnames:
            raise ValueError("CSV header is missing")
        headers = [self._header(value) for value in reader.fieldnames]
        rows = []
        for index, values in enumerate(reader, start=1):
            if index > self.MAX_ROWS:
                raise ValueError("Seller ERP source exceeds row limit")
            normalized = {
                headers[position]: self._cell(value)
                for position, value in enumerate(values.values())
                if position < len(headers)
            }
            if any(value for value in normalized.values()):
                rows.append(normalized)
        return headers, rows, None

    def _read_xlsx(
        self, content: bytes, worksheet: str | None
    ) -> tuple[list[str], list[dict[str, Any]], str]:
        try:
            workbook = load_workbook(
                io.BytesIO(content),
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            raise ValueError("XLSX workbook is invalid") from exc
        try:
            if worksheet:
                if worksheet not in workbook.sheetnames:
                    raise ValueError("XLSX worksheet is missing")
                sheet = workbook[worksheet]
            else:
                sheet = workbook[workbook.sheetnames[0]]
            iterator = sheet.iter_rows(values_only=True)
            try:
                first = next(iterator)
            except StopIteration as exc:
                raise ValueError("XLSX worksheet is empty") from exc
            headers = [self._header(value) for value in first]
            while headers and not headers[-1]:
                headers.pop()
            if not headers:
                raise ValueError("XLSX header is missing")
            rows = []
            for index, values in enumerate(iterator, start=1):
                if index > self.MAX_ROWS:
                    raise ValueError("Seller ERP source exceeds row limit")
                normalized = {
                    headers[position]: self._cell(value)
                    for position, value in enumerate(values[: len(headers)])
                }
                if any(value for value in normalized.values()):
                    rows.append(normalized)
            return headers, rows, sheet.title
        finally:
            workbook.close()

    def _normalize_row(
        self,
        *,
        domain: str,
        raw: dict[str, Any],
        column_map: dict[str, str],
        row_number: int,
    ) -> dict[str, Any]:
        values = {
            field: self._cell(raw.get(source_header))
            for field, source_header in column_map.items()
        }
        for required in self.SCHEMAS[domain]["required"]:
            if values.get(required, "") == "":
                raise ValueError(
                    f"row {row_number} missing required {required}"
                )
        if domain == "catalog":
            row = {
                "seller_sku": values["seller_sku"],
                "offer_id": values["offer_id"],
                "marketplace_sku": values.get("marketplace_sku", ""),
                "title": values["title"],
                "status": values["status"].lower(),
            }
        elif domain == "orders":
            row = {
                "order_external_id": values["order_external_id"],
                "seller_sku": values["seller_sku"],
                "status": values["status"].lower(),
                "quantity": self._integer(
                    values["quantity"], "quantity"
                ),
                "gross_revenue": self._decimal_text(
                    values["gross_revenue"], "gross_revenue"
                ),
                "currency": self._currency(values["currency"]),
                "updated_at": parse_timestamp(
                    values["updated_at"], "updated_at"
                ).isoformat(),
            }
        else:
            mode = values["fulfillment_mode"]
            if mode not in {"FBP", "realFBS"}:
                raise ValueError(
                    "fulfillment_mode must be FBP or realFBS"
                )
            row = {
                "seller_sku": values["seller_sku"],
                "warehouse_ref": values["warehouse_ref"],
                "fulfillment_mode": mode,
                **{
                    field: self._integer(values[field], field)
                    for field in (
                        "available_quantity",
                        "reserved_quantity",
                        "in_transit_quantity",
                        "damaged_quantity",
                        "quarantine_quantity",
                    )
                },
                "updated_at": parse_timestamp(
                    values["updated_at"], "updated_at"
                ).isoformat(),
            }
        row["canonical_key"] = self._key(domain, row)
        row["source_row_sha256"] = self._hash(
            {"row_number": row_number, "values": row}
        )
        return row

    def _column_map(
        self,
        *,
        domain: str,
        schema_version: str,
        column_map: Any,
    ) -> dict[str, str]:
        if domain not in self.SCHEMAS:
            raise ValueError("Unsupported Seller ERP domain")
        schema = self.SCHEMAS[domain]
        if schema_version != schema["version"]:
            raise ValueError("Seller ERP schema_version mismatch")
        if not isinstance(column_map, dict):
            raise ValueError("column_map must be an object")
        allowed = schema["required"] | schema["optional"]
        unknown = set(column_map) - allowed
        missing = schema["required"] - set(column_map)
        if unknown or missing:
            raise ValueError(
                "column_map fields do not match the versioned schema"
            )
        normalized = {
            self._required(str(key), "column_map field", 100): self._required(
                str(value), "column_map header", 300
            )
            for key, value in column_map.items()
        }
        if len(normalized.values()) != len(set(normalized.values())):
            raise ValueError("column_map source headers must be unique")
        return dict(sorted(normalized.items()))

    def _read_context(
        self,
        *,
        principal: Principal,
        entity_scope: dict[str, Any],
        store_ref: str,
        as_of: datetime,
    ) -> dict[str, Any]:
        if as_of.tzinfo is None or as_of.utcoffset() is None:
            raise ValueError("as_of must include a timezone")
        if not principal.can_access_store(store_ref):
            raise PermissionError(
                "Authenticated identity is not authorized for store_ref"
            )
        cutoff = as_of.astimezone(UTC)
        if entity_scope.get("status") != "ready":
            return {
                "status": "no_data",
                "cutoff": cutoff,
                "scope": {
                    "tenant_ref": principal.tenant_ref,
                    "entity_ref": None,
                    "store_ref": store_ref,
                    "scope_grant_authority_sha256": None,
                },
                "authority_sha256": None,
                "source_gaps": [
                    entity_scope.get(
                        "reason", "entity_scope_authority_missing"
                    )
                ],
            }
        entity_ref = str(entity_scope.get("entity_ref") or "").strip()
        authority = str(
            entity_scope.get("authority_sha256") or ""
        ).strip()
        if not entity_ref or len(authority) != 64:
            return {
                "status": "blocked",
                "cutoff": cutoff,
                "scope": {
                    "tenant_ref": principal.tenant_ref,
                    "entity_ref": None,
                    "store_ref": store_ref,
                    "scope_grant_authority_sha256": None,
                },
                "authority_sha256": None,
                "source_gaps": [
                    "seller_erp_bridge_entity_scope_authority_invalid"
                ],
            }
        return {
            "status": "ready",
            "cutoff": cutoff,
            "scope": {
                "tenant_ref": principal.tenant_ref,
                "entity_ref": entity_ref,
                "store_ref": store_ref,
                "scope_grant_authority_sha256": authority,
            },
            "authority_sha256": authority,
            "source_gaps": [],
        }

    def _write_context(
        self,
        *,
        principal: Principal,
        entity_scope: dict[str, Any],
        store_ref: str,
    ) -> dict[str, Any]:
        context = self._read_context(
            principal=principal,
            entity_scope=entity_scope,
            store_ref=store_ref,
            as_of=datetime.now(UTC),
        )
        if context["status"] != "ready":
            raise ValueError(context["source_gaps"][0])
        return context

    def _empty_result(
        self,
        *,
        context: dict[str, Any],
        source_evidence_id: str | None,
        page_size: int,
        cursor: str | None,
        query: str,
        state: str | None,
        gap: str | None = None,
    ) -> dict[str, Any]:
        gaps = sorted(
            set(context.get("source_gaps", []))
            | ({gap} if gap else set())
        )
        status = (
            context["status"]
            if context["status"] != "ready"
            else "no_data"
        )
        core = {
            "contract_id": self.CONTRACT_ID,
            "status": status,
            "as_of": context["cutoff"].isoformat(),
            "scope": context["scope"],
            "source": {
                "evidence_id": source_evidence_id,
                "sha256": None,
                "provider": None,
                "source_kind": None,
                "domain": None,
                "schema_version": None,
                "exported_at": None,
                "authorization_mode": None,
                "row_count": 0,
            },
            "authority": {
                "source_evidence_id": source_evidence_id,
                "source_evidence_sha256": None,
                "review_evidence_id": None,
                "binding_evidence_id": None,
                "revocation_evidence_id": None,
                "three_party_independence": False,
            },
            "query": {
                "page_size": page_size,
                "cursor": cursor,
                "next_cursor": None,
                "search": query or None,
                "state": state,
            },
            "counts": self._zero_counts(),
            "diff_items": [],
            "source_gaps": gaps,
            "blockers": [self._blocker(code) for code in gaps],
            "upstream_authority": {},
            "control_envelope": self._control(input_read=False),
        }
        return self._finish(core)

    def _blocked_result(
        self,
        *,
        context: dict[str, Any],
        source_evidence_id: str,
        page_size: int,
        cursor: str | None,
        query: str,
        state: str | None,
        gaps: list[str],
        authority: dict[str, Any],
        upstream: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        core = {
            "contract_id": self.CONTRACT_ID,
            "status": "blocked",
            "as_of": context["cutoff"].isoformat(),
            "scope": context["scope"],
            "source": {
                "evidence_id": source_evidence_id,
                "sha256": authority.get("source_evidence_sha256"),
                "provider": None,
                "source_kind": None,
                "domain": None,
                "schema_version": None,
                "exported_at": None,
                "authorization_mode": None,
                "row_count": 0,
            },
            "authority": authority,
            "query": {
                "page_size": page_size,
                "cursor": cursor,
                "next_cursor": None,
                "search": query or None,
                "state": state,
            },
            "counts": self._zero_counts(),
            "diff_items": [],
            "source_gaps": sorted(set(gaps)),
            "blockers": [
                self._blocker(code) for code in sorted(set(gaps))
            ],
            "upstream_authority": (
                {
                    "contract_id": upstream.get("contract_id"),
                    "status": upstream.get("status"),
                    "snapshot_sha256": upstream.get(
                        "snapshot_sha256"
                    ),
                }
                if upstream
                else {}
            ),
            "control_envelope": self._control(input_read=True),
        }
        return self._finish(core)

    def _finish(self, core: dict[str, Any]) -> dict[str, Any]:
        input_hash = self._hash(core)
        core["agent_artifact"] = {
            "contract_id": self.ARTIFACT_CONTRACT_ID,
            "input_snapshot_sha256": input_hash,
            "artifact_sha256": self._hash(
                {
                    "contract_id": self.ARTIFACT_CONTRACT_ID,
                    "input_snapshot_sha256": input_hash,
                    "suggestions": [],
                }
            ),
            "authority": "decision_support_and_internal_task_suggestion_only",
            "self_approval_allowed": False,
            "permit_issue_allowed": False,
            "formal_fact_promotion_allowed": False,
            "external_write_allowed": False,
        }
        core["snapshot_sha256"] = self._hash(core)
        return core

    @staticmethod
    def _zero_counts() -> dict[str, int]:
        return {
            "total_diff_items": 0,
            "page_diff_items": 0,
            "source_rows": 0,
            "canonical_rows": 0,
            "matched": 0,
            "source_only": 0,
            "canonical_only": 0,
            "conflict": 0,
            "blocked": 0,
        }

    @staticmethod
    def _control(*, input_read: bool) -> dict[str, bool]:
        return {
            "read_only": True,
            "scoped_input_read": input_read,
            "client_recalculation_allowed": False,
            "formal_fact_promoted": False,
            "product_created": False,
            "listing_created": False,
            "order_created": False,
            "inventory_created": False,
            "approval_created": False,
            "permit_created": False,
            "external_write_allowed": False,
            "private_interface_used": False,
        }

    @staticmethod
    def _blocker(code: str) -> dict[str, Any]:
        owner = (
            "identity-governance"
            if "entity_scope" in code
            else "integration-governance"
            if "schema" in code or "upstream" in code
            else "evidence-governance"
        )
        return {
            "code": code,
            "severity": "P0" if "invalid" in code or "revoked" in code else "P1",
            "owner": owner,
            "sla": "before reconciliation may inform an operating decision",
            "next": (
                "Repair exact entity/store authority."
                if owner == "identity-governance"
                else "Recapture the formal export with the reviewed schema mapping."
                if owner == "integration-governance"
                else "Complete independent review and compliance scope binding."
            ),
            "next_workspace": (
                "/scope-authority"
                if owner == "identity-governance"
                else "/seller-erp-bridge"
            ),
        }

    @classmethod
    def _source_projection(
        cls, record, *, status: str, idempotent: bool
    ) -> dict[str, Any]:
        return {
            "contract_id": cls.SOURCE_CONTRACT_ID,
            "status": status,
            "source_evidence_id": record.id,
            "source_evidence_sha256": record.sha256,
            "provider": record.metadata["provider"],
            "source_kind": record.metadata["source_kind"],
            "domain": record.metadata["domain"],
            "schema_version": record.metadata["schema_version"],
            "row_count": record.metadata["row_count"],
            "idempotent_replay": idempotent,
            "formal_fact_promoted": False,
            "external_write_allowed": False,
            "next": "Obtain an immutable independent review.",
        }

    @classmethod
    def _review_projection(
        cls, *, source, review, idempotent: bool
    ) -> dict[str, Any]:
        return {
            "contract_id": cls.REVIEW_CONTRACT_ID,
            "status": (
                "accepted_pending_compliance_binding"
                if review.metadata["decision"] == "accepted"
                else "rejected"
            ),
            "source_evidence_id": source.id,
            "source_evidence_sha256": source.sha256,
            "review_evidence_id": review.id,
            "review_evidence_sha256": review.sha256,
            "decision": review.metadata["decision"],
            "idempotent_replay": idempotent,
            "formal_fact_promoted": False,
            "external_write_allowed": False,
            "next": (
                "A separate compliance identity must record the exact-scope binding."
                if review.metadata["decision"] == "accepted"
                else "Recapture or correct the source; rejected Evidence remains immutable."
            ),
        }

    @classmethod
    def _binding_projection(
        cls, *, source, review, binding, idempotent: bool
    ) -> dict[str, Any]:
        return {
            "contract_id": cls.BINDING_WORKFLOW_CONTRACT_ID,
            "status": "bound_for_read_only_reconciliation",
            "source_evidence_id": source.id,
            "source_evidence_sha256": source.sha256,
            "review_evidence_id": review.id,
            "review_evidence_sha256": review.sha256,
            "binding_evidence_id": binding.id,
            "binding_evidence_sha256": binding.sha256,
            "three_party_independence": len(
                {source.created_by, review.created_by, binding.created_by}
            )
            == 3,
            "idempotent_replay": idempotent,
            "formal_fact_promoted": False,
            "external_write_allowed": False,
            "next": "Run read-only Canonical Diff.",
        }

    @classmethod
    def _revocation_projection(
        cls, *, source, revocation, idempotent: bool
    ) -> dict[str, Any]:
        return {
            "contract_id": cls.REVOCATION_CONTRACT_ID,
            "status": "revoked",
            "source_evidence_id": source.id,
            "source_evidence_sha256": source.sha256,
            "revocation_evidence_id": revocation.id,
            "revocation_evidence_sha256": revocation.sha256,
            "idempotent_replay": idempotent,
            "formal_fact_promoted": False,
            "external_write_allowed": False,
            "next": "Capture and independently authorize a new source snapshot.",
        }

    @classmethod
    def _unique_rows(
        cls, rows: list[dict[str, Any]]
    ) -> list[dict[str, Any]]:
        seen = set()
        ordered = []
        for row in rows:
            key = row.get("canonical_key")
            if not key:
                raise ValueError("canonical key is missing")
            if key in seen:
                raise ValueError(
                    f"duplicate canonical key in one snapshot: {key}"
                )
            seen.add(key)
            ordered.append(row)
        ordered.sort(key=lambda item: item["canonical_key"])
        return ordered

    @staticmethod
    def _key(domain: str, row: dict[str, Any]) -> str:
        if domain == "catalog":
            fields = (row.get("seller_sku"), row.get("offer_id"))
        elif domain == "orders":
            fields = (row.get("order_external_id"),)
        else:
            fields = (
                row.get("seller_sku"),
                row.get("warehouse_ref"),
                row.get("fulfillment_mode"),
            )
        values = [str(value or "").strip() for value in fields]
        if any(not value for value in values):
            raise ValueError(f"{domain} canonical key is incomplete")
        return "|".join(values)

    @classmethod
    def _integer(cls, value: Any, field: str) -> int:
        try:
            number = Decimal(str(value))
        except InvalidOperation as exc:
            raise ValueError(f"{field} must be a non-negative integer") from exc
        if not number.is_finite() or number < 0 or number != number.to_integral_value():
            raise ValueError(f"{field} must be a non-negative integer")
        return int(number)

    @staticmethod
    def _decimal_text(value: Any, field: str) -> str:
        try:
            number = Decimal(str(value))
        except InvalidOperation as exc:
            raise ValueError(f"{field} must be a non-negative decimal") from exc
        if not number.is_finite() or number < 0:
            raise ValueError(f"{field} must be a non-negative decimal")
        return format(number, "f")

    @staticmethod
    def _currency(value: str) -> str:
        currency = value.strip().upper()
        if (
            len(currency) != 3
            or not currency.isascii()
            or not currency.isalpha()
        ):
            raise ValueError("currency must be a three-letter ASCII code")
        return currency

    @classmethod
    def _header(cls, value: Any) -> str:
        header = cls._cell(value)
        if not header:
            raise ValueError("Seller ERP source contains an empty header")
        return header

    @classmethod
    def _cell(cls, value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        if len(text) > cls.MAX_CELL_CHARS:
            raise ValueError("Seller ERP source cell exceeds size limit")
        if "\x00" in text:
            raise ValueError("Seller ERP source cell contains NUL")
        return text

    @staticmethod
    def _require_scope(
        metadata: dict[str, Any], expected: dict[str, Any]
    ) -> None:
        for key in ("tenant_ref", "entity_ref", "store_ref"):
            if metadata.get(key) != expected.get(key):
                raise ValueError(f"{key} mismatch")

    @staticmethod
    def _required(value: str | None, field: str, limit: int) -> str:
        normalized = str(value or "").strip()
        if not normalized or len(normalized) > limit:
            raise ValueError(f"{field} is required and must be <= {limit}")
        return normalized

    @classmethod
    def _choice(
        cls, value: Any, field: str, allowed: frozenset[str]
    ) -> str:
        normalized = cls._required(str(value or ""), field, 100)
        if normalized not in allowed:
            raise ValueError(f"Unsupported {field}: {normalized}")
        return normalized

    @classmethod
    def _require_replay(
        cls,
        *,
        existing,
        content: bytes,
        effective_at: datetime,
        metadata: dict[str, Any],
        created_by: str,
    ) -> None:
        if (
            existing.sha256 != cls._sha(content)
            or cls._stored_time(existing.effective_at) != effective_at
            or existing.metadata != metadata
            or existing.created_by != created_by
        ):
            raise ValueError(
                "Seller ERP bridge idempotency key conflicts with immutable request"
            )

    @staticmethod
    def _record_key(record) -> tuple[datetime, datetime, str]:
        return (
            ScopedSellerErpBridge._stored_time(record.effective_at),
            ScopedSellerErpBridge._stored_time(record.recorded_at),
            record.id,
        )

    @staticmethod
    def _stored_time(value: str) -> datetime:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=UTC)
        return parsed.astimezone(UTC)

    @classmethod
    def _safe_gap(cls, prefix: str, exc: Exception) -> str:
        text = str(exc).casefold()
        suffix = (
            "scope_conflict"
            if "scope" in text or "tenant" in text or "store" in text
            else "hash_conflict"
            if "hash" in text
            else "contract_conflict"
            if "contract" in text
            else "latest_rejected"
            if "not accepted" in text
            else "invalid"
        )
        return f"{prefix}:{suffix}"

    @staticmethod
    def _encode_cursor(key: tuple[str, str]) -> str:
        payload = json.dumps(
            {"state": key[0], "canonical_key": key[1]},
            sort_keys=True,
            separators=(",", ":"),
        ).encode()
        return base64.urlsafe_b64encode(payload).decode().rstrip("=")

    @staticmethod
    def _decode_cursor(value: str) -> tuple[str, str]:
        try:
            padded = value + "=" * (-len(value) % 4)
            decoded = json.loads(
                base64.urlsafe_b64decode(padded).decode()
            )
            state = str(decoded["state"])
            key = str(decoded["canonical_key"])
        except (
            binascii.Error,
            KeyError,
            UnicodeDecodeError,
            json.JSONDecodeError,
            TypeError,
            ValueError,
        ) as exc:
            raise ValueError("Invalid Seller ERP bridge cursor") from exc
        if not state or not key:
            raise ValueError("Invalid Seller ERP bridge cursor")
        return state, key

    @staticmethod
    def _json_bytes(value: Any) -> bytes:
        return json.dumps(
            value,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode()

    @classmethod
    def _hash(cls, value: Any) -> str:
        return cls._sha(cls._json_bytes(value))

    @staticmethod
    def _sha(content: bytes) -> str:
        return hashlib.sha256(content).hexdigest()
